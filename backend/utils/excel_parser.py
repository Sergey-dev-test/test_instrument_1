# backend/utils/excel_parser.py

from openpyxl import load_workbook
from pandas import DataFrame
from typing import Dict, List, Any, Tuple, Optional


def validate_excel_column(data: List[Any], column_name: str, expected_type: type, is_required: bool = True) -> Tuple[bool, Optional[str]]:
    """
    Проверка колонки Excel:
    - Заполнено ли значение (если required)
    - Соответствует ли тип
    """
    for i, value in enumerate(data, start=2):  # i=2, т.к. строка 1 — заголовок
        if value is None or value == "":
            if is_required:
                return False, f"Строка {i}: поле '{column_name}' пустое (обязательно)"
        else:
            if not isinstance(value, expected_type):
                return False, f"Строка {i}: неверный тип в '{column_name}' (ожидался {expected_type.__name__})"
    return True, None


def parse_excel_with_validation(file_path: str, expected_columns: Dict[str, type], required_columns: List[str]) -> Tuple[bool, DataFrame, List[str]]:
    """
    Парсинг Excel + валидация:
    - expected_columns: {"username": str, "age": int}
    - required_columns: ["username"]
    Возвращает:
    - is_valid: bool
    - df: DataFrame (если валидно)
    - errors: List[str]
    """
    errors = []
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Проверка заголовков
        headers = [cell.value for cell in ws[1]]
        for col in expected_columns.keys():
            if col not in headers:
                errors.append(f"Отсутствует обязательный столбец: '{col}'")
        
        # Сбор данных
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is None for cell in row):  # Пропуск пустых строк
                continue
            data.append(row)

        df = DataFrame(data, columns=headers[:len(expected_columns)])

        # Валидация типов и обязательных полей
        for col, col_type in expected_columns.items():
            is_required = col in required_columns
            is_valid_col, err = validate_excel_column(df[col].tolist(), col, col_type, is_required)
            if not is_valid_col:
                errors.append(err)

        return len(errors) == 0, df, errors
    except Exception as e:
        return False, DataFrame(), [f"Ошибка чтения файла: {str(e)}"]